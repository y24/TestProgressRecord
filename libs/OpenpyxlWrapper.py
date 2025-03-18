from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

def load(file_path:str, auto_create:bool=False):
    # ブックを開く
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        # ファイルが存在しない場合、新規作成
        if auto_create:
            wb = Workbook()
        else:
            raise FileNotFoundError(f"Error: {file_path} が見つかりません。")
    except PermissionError:
        raise PermissionError(f"Error: '{file_path}' は他のプログラムによって開かれています。")
    return wb

def create_sheet(workbook, sheet_name:str, overwrite:bool=False):
    # 既存のデータシートがあれば削除
    if overwrite and sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    # 新しいデータシートを作成
    return workbook.create_sheet(title=sheet_name)


def get_sheet_by_name(workbook, sheet_name:str):
    return workbook[sheet_name]


def get_sheetnames_by_keyword(workbook, keyword:str):
    return [sheet for sheet in workbook.sheetnames if keyword in sheet]

def get_sheetnames_by_keywords(workbook, keywords: list, ignores: list = None):
    if ignores is None:
        ignores = []
    return [
        sheet for sheet in workbook.sheetnames 
        if any(keyword in sheet for keyword in keywords) and not any(ignore in sheet for ignore in ignores)
    ]


def find_row(sheet, search_col:str, search_str:str):
    try:
        # 列名
        col_num = column_index_from_string(search_col) - 1

        # 指定列をループして値を確認
        for row in sheet.iter_rows(min_col=1, max_col=1):
            cell = row[col_num]  # 指定列のセル
            if cell.value == search_str:  # 値が search_str のセル
                return cell.row
        return None

    except Exception as e:
        print(f"Error: {e}")


def get_row_values(sheet, row_num:int):
    return [cell.value for cell in sheet[row_num]]


def get_column_values(sheet, col_nums: list, header_row: int = 1, ignore_header=False):
    if ignore_header:
        header_row += 1
    return [[sheet.cell(row=i, column=col_num).value for col_num in col_nums] for i in range(header_row, sheet.max_row + 1)]


def get_columns_data(sheet, col_nums: list, header_row: int = 1, ignore_header=False):
    if ignore_header:
        header_row += 1
    data = [[cell.value.strftime('%Y-%m-%d') if isinstance(cell.value, datetime) else cell.value 
             for col_num in col_nums 
             for cell in [sheet.cell(row=i, column=col_num)]] 
            for i in range(header_row, sheet.max_row + 1)]
    return data

def get_cell_value(sheet, col:int, row:int, replace_newline=False):
    value = sheet.cell(row=row, column=col).value
    if replace_newline:
        if value and isinstance(value, str):  # 値が文字列の場合のみ変換
            return value.replace("\n", "_")
    else:
        return value
